import openpyxl
from datetime import datetime, timedelta # timedelta는 날짜를 더하고 빼기


# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 현재 활성화된 시트 선택
# ws = wb.active

# 새로운 시트 생성
wss = wb.create_sheet('2020.03')
# 시트 삭제
del wb['Sheet']

# 시트이름 변경
wss.title = 'automation'

# 경로 
path = 'C:/Users/yongwon/Jupyter/automation.xlsx'

# 엑셀 저장
wb.save(path)

# 엑셀 파일 불러오기
openpyxl.load_workbook(path)

# 활성화된 시트 선택
ws = wb.active

# 데이터 추가(1)
ws['A1'] = '날짜'
ws['B1'] = '가격'
ws['C1'] = '제품명'
ws['D1'] = '수량'
ws['E1'] = '합계'

# 날짜 시작
start_date = datetime(2030, 1, 1)

# 데이터 추가(2)
ws.cell(row=2, column=1, value= start_date.strftime('%Y-%m-%d'))
ws.cell(row=2, column=2, value = 20000)
ws.cell(row=2, column=3, value='삼성tv')
ws.cell(row=2, column=4, value = 4)
ws.cell(row=2, column=5, value='=B2*D2')

for i in range(3,7):
    current_date  = start_date + timedelta(days=i-2)
    ws.cell(row = i, column = 1, value = current_date.strftime('%Y-%m-%d'))

# 데이터 추가(3)
ws.append(['2030-01-01', 1200000,'기계식키보드', 15, '=B5*D5'])

# 데이터 수정
ws['C2'] = 'LGtv'
del ws['A3']

wb.save(path)