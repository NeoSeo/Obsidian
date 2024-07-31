import openpyxl
import pandas as pd

save_path1 = 'C:/Users/yongwon/Jupyter/item_selling (3).xlsx'
save_path2 = 'C:/Users/yongwon/Jupyter/item_selling (4).xlsx'
save_path3 = 'C:/Users/yongwon/Jupyter/item_selling (5).xlsx'

save_path = [save_path1, save_path2, save_path3]

# wb = openpyxl.Workbook()
# total_ws = wb.active
data = []

for i in save_path:
    wb = openpyxl.load_workbook(i, data_only = True) ## 엑셀 수식을 무시하고 data만 가져오는 옵션
    ws = wb['automation']
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=5, values_only=True): #max_row = ws.max_row 필요없다
        data.append(list(row))
    # ws = wb.active
    # for row in ws.iter_rows(min_row = 2):
    #   for cell in row:
    #       data.append(cell.value)
    #   total_ws.append(data)
        
path = 'C:/Users/yongwon/Jupyter/total.xlsx'
df = pd.DataFrame(data, columns = ['제품명','가격','수량','합계'])
df.insert(0, '순번', range(1, len(df) + 1))

df.to_excel(path, index = False)