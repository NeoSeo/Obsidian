import openpyxl

save_path = 'C:/Users/yongwon/Jupyter/item_selling (3).xlsx'

wb = openpyxl.load_workbook(save_path, data_only = True) # data_only는 수식 무시하고 데이터만 가져옮

ws = wb['automation']

# 행과 열을 개수를 알 경우
for x in range(1,12):
    for y in range(1, 6):
        print(ws.cell(row=x, column=y).value, end=" ") # end는 한칸 띄우기
    print() # 마지막 줄바꿈

print()
# 행과 열 개수를 모를 경우    
for x in range(1,ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=" ") # end는 한칸 띄우기
    print()

print()

# 모든 행 가져오기
for row in ws.iter_rows():
    print(row)

print()
# 두번 째 행부터 5째 행까지, 2~4 열까지
for row in ws.iter_rows(min_row=2, max_row=5, min_col=2, max_col=4):
    for cell in row:
       print(cell.value, end=" ")
    print()
