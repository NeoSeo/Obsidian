import openpyxl
import random
import os
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active

base_path = 'C:/Users/yongwon/Jupyter'
base_filename = 'item_selling.xlsx'
full_path = os.path.join(base_path, base_filename)

# 파일명이 겹치면 숫자를 붙여가며 저장
if os.path.exists(full_path):
    version = 1
    while os.path.exists(full_path):
        name, ext = os.path.splitext(base_filename)
        full_path = os.path.join(base_path, f"{name} ({version}){ext}")
        version += 1

ws.title = 'automation'

item = [['32인치 모니터', 350000], ['마우스 패드', 20000],['게이밍 마우스', 40000], ['기계식 키보드',120000]]

# ws['A1'] = '순번'
# ws['B1'] = '제품명'
# ws['C1'] = '가격'
# ws['D1'] = '수량'
# ws['E1'] = '합계'
# ws['F1'] = '날짜'

ws.append(['순번','제품명','가격','수량','합계','날짜'])

first_date = datetime(2024, 3, 1)

for i in range(2, 12):
    current_date = first_date + timedelta(i)
    rand_ind = random.randint(0,3)
    selected_item = item[rand_ind]
    ws.append([i-1, selected_item[0], selected_item[1], random.randint(1,8), f'=C{i}*D{i}', current_date.strftime('%Y-%m-%d')]) # strftime으로 나타내야 엑셀에서 날짜형식으로 뜬다.

wb.save(full_path)

